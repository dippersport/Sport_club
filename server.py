# import tkinter as tk
# from tkinter import simpledialog
# import xlsxwriter

# # Функция для открытия модального окна с формой анкеты
# def open_modal():
#     # Создаем диалоговое окно
#     root = tk.Tk()
#     root.withdraw()  # скрываем основное окно

#     # Запрашиваем данные у пользователя
#     name = simpledialog.askstring("Анкета участника", "Введите ФИО:")
#     age = simpledialog.askinteger("Анкета участника", "Дата рождения:")
#     section = simpledialog.askstring("Анкета участника", "Введите Адрес проживания:")
#     medical = simpledialog.askstring("Анкета участника", "Введите медицинскую справку:")
#     classroom = simpledialog.askstring("Анкета участника", "Введите Класс:")
#     school = simpledialog.askstring("Анкета участника", "Введите Название школы:")

#     parent_consent = simpledialog.askstring("Анкета участника", "Введите разрешение от родителей:")

#     # Создаем или открываем файл Excel
#     workbook = xlsxwriter.Workbook('anketa.xlsx')
#     worksheet = workbook.add_worksheet()

#     # Записываем данные в файл Excel
#     row = 0
#     worksheet.write(row, 0, "ФИО")
#     worksheet.write(row, 1, "Возраст")
#     worksheet.write(row, 2, "Секция")
#     worksheet.write(row, 3, "Медицинская справка")
#     worksheet.write(row, 4, "Класс")
#     worksheet.write(row, 5, "Название школы:")
#     worksheet.write(row, 6, "Разрешение от родителей")
#     row += 1
#     worksheet.write(row, 0, name)
#     worksheet.write(row, 1, age)
#     worksheet.write(row, 2, section)
#     worksheet.write(row, 3, medical)
#     worksheet.write(row, 4, classroom)
#     worksheet.write(row, 5, school)
#     worksheet.write(row, 6, parent_consent)

#     # Закрываем файл Excel
#     workbook.close()

# # Вызываем функцию для открытия модального окна
# open_modal()


from flask import Flask, request, jsonify
import xlsxwriter
import os
from openpyxl import load_workbook

app = Flask(__name__)

@app.route('/submit', methods=['POST'])
def submit():
    name = request.form.get('name')
    age = request.form.get('age')
    address = request.form.get('address')
    medical = request.form.get('medical')
    classroom = request.form.get('class')
    school = request.form.get('school')
    parent_consent = request.form.get('parent-consent')

    file_path = 'anketa.xlsx'
    file_exists = os.path.exists(file_path)

    if file_exists:
        # Если файл существует, открываем его для добавления данных
        workbook = load_workbook(file_path)
        worksheet = workbook.active
        row = worksheet.max_row + 1
    else:
        # Если файл не существует, создаем новый
        workbook = xlsxwriter.Workbook(file_path)
        worksheet = workbook.add_worksheet()
        row = 0
        # Пишем заголовки в новой таблице
        worksheet.write(row, 0, "ФИО")
        worksheet.write(row, 1, "Дата рождения")
        worksheet.write(row, 2, "Адрес проживания")
        worksheet.write(row, 3, "Медицинская справка")
        worksheet.write(row, 4, "Класс")
        worksheet.write(row, 5, "Название школы")
        worksheet.write(row, 6, "Разрешение от родителей")
        row += 1

    # Записываем данные в файл Excel
    worksheet.write(row, 0, name)
    worksheet.write(row, 1, age)
    worksheet.write(row, 2, address)
    worksheet.write(row, 3, medical)
    worksheet.write(row, 4, classroom)
    worksheet.write(row, 5, school)
    worksheet.write(row, 6, parent_consent)

    # Закрываем файл Excel
    workbook.close()

    return jsonify({"message": "Данные успешно сохранены!"})

if __name__ == '__main__':
    app.run(debug=True)

